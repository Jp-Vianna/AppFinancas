import openpyxl


class Planilha:
    def __init__(self):
        self.book = openpyxl.load_workbook('PlanilhaDespezas.xlsx')     # Abre planilha.
        self.pagina_despezas = self.book["Despezas"]                    # Seleciona página a ser trabalhada.

    def adiciona_linha(self, valor, titulo, tipo):
        self.pagina_despezas.append([valor, titulo, tipo])      # Adiciona nova linha.
        self.book.save('PlanilhaDespezas.xlsx')                 # Salva mudanças do arquivo.

    # Remove uma linha da planilha.
    def remove(self, despeza):
        contador = 1        # Conta as linhas percorridas.
        for linha in self.pagina_despezas.iter_rows(min_row=2):     # Percorre todas as linhas do arquivo.
            contador += 1       # Incrementa.

            if self.verifica_vazio(linha):      # Se a linha é vazia, para.
                return

            if self.verifica_igualdade(linha, despeza.valor, despeza.titulo, despeza.tipo):         # Se a linha é a procurada, exclui.
                self.pagina_despezas.delete_rows(contador, 1)                             # Exclui a linha do index atual(contador).
                self.book.save("PlanilhaDespezas.xlsx")                                        # Salva modificações.

                break

    # Método que procura uma linha para substituir pelos novos valores.
    def edita(self, despeza, valores):
        for linha in self.pagina_despezas.iter_rows(min_row=2):
            if self.verifica_vazio(linha):
                return

            if self.verifica_igualdade(linha, valores[0], valores[1], valores[2]):      # Se for a linha procurada, substitui.
                linha[0].value = f"R$ {despeza.valor:.2f}"              # Substitui todos os valores.
                linha[1].value = despeza.titulo
                linha[2].value = despeza.tipo

                self.book.save("PlanilhaDespezas.xlsx")

                break

    # Método estático que verifica se a linha da planilha é igual aos dados procurados.
    @staticmethod
    def verifica_igualdade(linha, valor, titulo, tipo):
        return (float(linha[0].value[3:]) == valor) and (linha[1].value == titulo) and (linha[2].value == tipo)

    # Método estático que verifica se alinha está vazia.
    @staticmethod
    def verifica_vazio(linha):
        return any(celulas.value is None for celulas in linha[:3])
